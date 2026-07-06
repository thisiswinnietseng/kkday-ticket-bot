import asyncio, os, re, json, threading, io
from datetime import datetime, timedelta
import requests as http_requests
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright
from dotenv import load_dotenv

load_dotenv(os.path.expanduser('~/kkday-ticket-bot/.env'))

class StopRequested(Exception):
    pass

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)
BE2_USERNAME = os.environ.get('BE2_USERNAME', '')
BE2_PASSWORD = os.environ.get('BE2_PASSWORD', '')
HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'history.json')
SHEETS_WEBHOOK = os.environ.get('SHEETS_WEBHOOK', '')

# ── 額滿挽單 對客通知模板（依語系 × 備註類型）──────────────
NOTIF_SUBJECT = '額滿挽單 Sold-out Booking Recovery'

NOTIF_TEMPLATES = {
    'page': {
        'zh-tw': """您好，感謝您選擇 KKday 作為您的旅途夥伴。😊

很抱歉，收到供應商通知，您訂購的日期/時間已經額滿，請問您是否考慮更改日期/時間體驗呢？
現可更改的 日期/時間 請依照商品頁面提供為主。

請提供您的意願順序，

第一候補場次：______________
第二候補場次：______________
（將自動依您的排序預訂，不再另行通知）

目前訂單尚未成功，煩請您於【 D/L 】(GMT+8) 以前回覆。若在這之前都沒有收到您的回覆，將無法完成您的訂單且會取消。
若以上日期都沒有您想要的，那麼您可以回覆告知您想要的日期讓我們和相關部門查看或回覆“取消預定”，我們會再為您申請。
若您還有其他需要協助的地方，歡迎您再來訊，謝謝您。😊

期待您的回覆。🙏""",
        'zh-cn': """您好，感谢您选择 KKday 作为您的旅途伙伴。😊

很抱歉，收到供应商通知，您订购的日期/时间已经额满，请问您是否考虑更改日期/时间体验呢？
现可更改的 日期/时间 请依照商品页面提供为主。

请提供您的意愿顺序，

第一候补场次：______________
第二候补场次：______________
（将自动依您的排序预订，不再另行通知）

目前订单尚未成功，烦请您于 【D/L 】 (GMT+8) 以前回复。若在这之前都没有收到您的回复，将无法完成您的订单且会取消。
若以上日期都没有您想要的，那麽您可以回复告知您想要的日期让我们和相关部门查看或回复“取消预定”，我们会再为您申请。
若您还有其他需要协助的地方，欢迎您再来讯，谢谢您。😊""",
        'zh-hk': """您好，感謝您選擇 KKday 作為您的旅途夥伴。😊

很抱歉，收到供應商通知，您訂購的日期/時間已經額滿，請問您是否考慮更改日期/時間體驗呢？
現可更改的 日期/時間 請依照商品頁面提供為主。

請提供您的意願順序，

第一候補場次：______________
第二候補場次：______________
（將自動依您的排序預訂，不再另行通知）

目前訂單尚未成功，煩請您於 【D/L  】 (GMT+8) 以前回覆。若在這之前都沒有收到您的回覆，將無法完成您的訂單且會取消。
若以上日期都沒有您想要的，那麼您可以回覆告知您想要的日期讓我們和相關部門查看或回覆“取消預定”，我們會再為您申請。
若您還有其他需要協助的地方，歡迎您再來訊，謝謝您。

期待您的回覆。🙏""",
        'en': """Hello, valued KKday traveler, thank you for choosing KKday and we appreciate your waiting.

Unfortunately, we received a notice from the supplier that your selected time slot is fully booked.

Do you wish to amend your trip time? 
Time available for amendment is shown on the product page.
(On a first come first serve basis, not reserved)

You may also provide first priority or backup dates for better arrangement, thank you.
First priority: _________
Back-up dates: _________

If you're not available for the amendment, we can also help with canceling the booking. 
Sorry for any inconvenience caused.

Please kindly get back to us before 【D/L  】(GMT+8), or your booking will be forfeited and we will assist to cancel and full refund for you, 

Seeking your confirmation, Thank you.""",
        'ko': """안녕하십니까 고객님, 저희 KKday를 이용해 주셔서 감사합니다.😇

고객님의 예약이 확인 되었으나, 해당 상품이 예약 조기 마감으로 상품 제공이 어렵게 되었음을 안내드립니다.😥
예약 가능날짜 및 시간  상품페이지에 확인 후 변경 원하시는 날짜 및 시간  제공부탁드립니다.

*본 상품은 실시간으로 예약이 진행되어, 간혹 날짜 및 시간 조기 마감이 될 수 있으니 이 점 양해 부탁드립니다.

①변경 원하시는날짜 및 시간 : ______________
②후보시간 제공
후보 : ______________
(자동으로 고객님 순서에 따라 예약되며, 추후 공지하지 않습니다)

③시간 변경 또는 예약 취소를 희망하실 경우 회신 부탁드립니다.


현재 예약이 아직 완료되지 않았으니 【D/L  】 (GMT+8)까지 회신 부탁드립니다. 답변을 받지 못한 경우 예약은 취소드리는 점 양해부탁드립니다..

다른 문의사항이 있으시면 언제든지 KKday 고객센터로 문의 주시길 바랍니다. 😊
감사합니다. 즐거운 자유여행은 KKday와 함께하세요. ：）✈️""",
        'ja': """KKdayをご利用いただきありがとうございます。

恐れ入りますがご予約いただきました日程は完売のため手配ができませんでした。

つきまして
A：商品ページよりご選択いただける日程へ変更を希望（ご希望順をお知らせください）
　※現時点での情報であり、空席を保証するものではございません。
B：都合がつかないのでキャンセル希望
どちらかをご選択の上、お忙しい中恐れ入りますが台湾時間【D/L  】(GMT+8)までにご返信いただきますようよろしくお願いいたします。""",
    },
    'op': {
        'zh-tw': """您好，感謝您選擇 KKday 作為您的旅途夥伴。😊

很抱歉，收到供應商通知，您訂購的日期/時間已經額滿，請問您是否考慮更改日期/時間體驗呢？現可更改為：

👉【 】
（名額不保留，先搶先得）

請提供您的意願順序，

第一候補場次：______________
第二候補場次：______________
（將自動依您的排序預訂，不再另行通知）

目前訂單尚未成功，煩請您於 【D/L  】  (GMT+8) 以前回覆。若在這之前都沒有收到您的回覆，將無法完成您的訂單且會取消。
若以上日期都沒有您想要的，那麼您可以回覆告知您想要的日期讓我們和相關部門查看或回覆“取消預定”，我們會再為您申請。
若您還有其他需要協助的地方，歡迎您再來訊，謝謝您。😊

期待您的回覆。🙏""",
        'zh-cn': """您好，感谢您选择 KKday 作为您的旅途伙伴。😊

很抱歉，收到供应商通知，您订购的日期/时间已经额满，请问您是否考虑更改日期/时间体验呢？现可更改为：

👉【 】
（名额不保留，先抢先得）

请提供您的意愿顺序，

第一候补场次：______________
第二候补场次：______________
（将自动依您的排序预订，不再另行通知）

目前订单尚未成功，烦请您于 【D/L  】 (GMT+8) 以前回复。若在这之前都没有收到您的回复，将无法完成您的订单且会取消。
若以上日期都没有您想要的，那麽您可以回复告知您想要的日期让我们和相关部门查看或回复“取消预定”，我们会再为您申请。
若您还有其他需要协助的地方，欢迎您再来讯，谢谢您。😊

期待您的回复。🙏""",
        'zh-hk': """您好，感謝您選擇 KKday 作為您的旅途夥伴。😊

很抱歉，收到供應商通知，您訂購的日期/時間已經額滿，請問您是否考慮更改日期/時間體驗呢？現可更改為：

👉【 】
（名額不保留，先搶先得）

請提供您的意願順序，

第一候補場次：______________
第二候補場次：______________
（將自動依您的排序預訂，不再另行通知）

目前訂單尚未成功，煩請您於 【D/L  】 (GMT+8) 以前回覆。若在這之前都沒有收到您的回覆，將無法完成您的訂單且會取消。
若以上日期都沒有您想要的，那麼您可以回覆告知您想要的日期讓我們和相關部門查看或回覆“取消預定”，我們會再為您申請。
若您還有其他需要協助的地方，歡迎您再來訊，謝謝您。

期待您的回覆。🙏""",
        'en': """Hello, valued KKday traveler, thank you for choosing KKday and we appreciate your waiting.

Unfortunately, we received a notice from the supplier that your selected time slot is fully booked.

If you wish to reschedule, the time slots available for amendment currently are:

👉【 】
(On a first come first serve basis, not reserved)

Please provide your priority according the above information for better arrangement, thank you.

First priority: _________
Back-up dates: _________

If you're not available for the amendment, we can also help with canceling the booking as well, sorry for any inconvenience caused.

Please kindly get back to us before 【D/L  】(GMT+8), or your booking will be forfeited and we will assist to cancel and full refund for you, 

Seeking your confirmation, Thank you.""",
        'ko': """안녕하십니까 고객님, 저희 KKday를 이용해 주셔서 감사합니다.😇

고객님의 예약이 확인 되었으나, 해당 상품이 예약 조기 마감으로 상품 제공이 어렵게 되었음을 안내드립니다.😥
현재 하기의 날짜 및 시간 예약이 가능한 점 참고 부탁드립니다. 

👉【 】

*본 상품은 실시간으로 예약이 진행되어, 간혹 시간대가 조기 마감이 될 수 있으니 이 점 양해 부탁드립니다.


1차 후보 : _______________
2차 후보 : _______________
(자동으로 고객님 순서에 따라 예약되며, 추후 공지하지 않습니다)

시간 변경 또는 예약 취소를 희망하실 경우 회신 부탁드립니다.

현재 예약이 아직 완료되지 않았으니 【D/L  】 (GMT+8)까지 회신 부탁드립니다. 답변을 받지 못한 경우 예약은 취소드리는 점 양해부탁드립니다..

다른 문의사항이 있으시면 언제든지 KKday 고객센터로 문의 주시길 바랍니다. 😊
감사합니다. 즐거운 자유여행은 KKday와 함께하세요. ：）✈️""",
        'ja': """KKdayをご利用いただきありがとうございます。

恐れ入りますがご予約いただきました日程は完売のため手配ができませんでした。
現在以下の日程にて空きがある状況でございます。
‧ 【 】

※現時点での情報であり、空席を保証するものではございません。

つきまして
A：上記いずれかの日程へ変更を希望（ご希望順をお知らせください）
B：都合がつかないのでキャンセル希望
どちらかをご選択の上、お忙しい中恐れ入りますが台湾時間【D/L  】(GMT+8)までにご返信いただきますようよろしくお願いいたします。""",
    },
}

# ── 未成團挽單 對客通知模板（依語系）──────────────────────────
NOTIF_SUBJECT_MITUAN = '未成團挽單 Tour Cancellation Recovery'

NOTIF_TEMPLATES_MITUAN = {
    'zh-tw': """您好，感謝您的訂購😊

很遺憾接獲供應商告知您所訂購行程因 未達最低成團人數而無法成團，
目前可改為 👉【  】
（名額不保留，先搶先贏）請您確認您是否能接受此安排，以利為您後續作業。

①我要更改的日期／場次：______________
②若有想候補的備選時間，亦請您提供以下資訊👇🏻
第一候補場次：______________
(將自動依您的排序預訂，不再另行通知)
③若您無法配合以上日期，請您直接回覆我們：「我要取消訂單」，將為您取消訂單並全額退費。

煩請您於【D/L  】(GMT+8) 以前回覆，若逾時未回覆，將為您辦理取消並全額退款。造成您的不便，望您諒解，謝謝您。""",
    'zh-cn': """您好，感谢您选择KKday作为您的旅途伙伴。😊

很遗憾接获供应商告知您所订购行程因 未达最低成团人数而无法成团，
目前可改为 👉【  】
（名额不保留，先抢先赢）请您确认您是否能接受此安排，以利为您后续作业。

①我要更改的日期／场次：______________
②若有想候补的备选时间，亦请您提供以下资讯👇🏻
第一候补场次：______________
(将自动依您的排序预订，不再另行通知)
③若您无法配合以上日期，请您直接回复我们：「我要取消订单」，将为您取消订单并全额退费。

烦请您于【D/L  】(GMT+8) 以前回复，若逾时未回复，将为您办理取消并全额退款。造成您的不便，望您谅解，谢谢您。""",
    'zh-hk': """您好，感謝您選擇KKday作為您的旅途夥伴。😊

很遺憾接獲供應商告知您所訂購行程因 未達最低成團人數而無法成團，
目前可改為 👉【  】
（名額不保留，先搶先贏）請您確認您是否能接受此安排，以利為您後續作業。

①我要更改的日期／場次：______________
②若有想候補的備選時間，亦請您提供以下資訊👇🏻
第一候補場次：______________
(將自動依您的排序預訂，不再另行通知)
③若您無法配合以上日期，請您直接回覆我們：「我要取消訂單」，將為您取消訂單並全額退費。

煩請您於【D/L 】(GMT+8) 以前回覆，若逾時未回覆，將為您辦理取消並全額退款。造成您的不便，望您諒解，謝謝您。🙏""",
    'en': """Hello, valued KKday traveler, thank you for choosing KKday and we appreciate your waiting.

We are sorry to inform you that the tour date you selected haven't met the required travelers, so we have no choice but to cancel the reservation.

Currently the available time slots are as below:
👉【  】
(On first come first serve basis, not reserved)

Please confirm whether you can accept this amendment, for further follow-up,

①The date/session you want to change: ______________

②If you have another desired time slot, please also provide the following information
(will be automatically booked according to your order without further notice)
First alternate session: ______________

③If you wish to cancel instead of amending the tour, please reply us directly: "I want to cancel the order", the order will be canceled for you and a full refund will be made.

Please kindly get back to us before 【D/L  】(GMT+8), or your booking will be forfeited and we will assist to cancel and full refund for you,

Seeking your confirmation, Thank you.""",
    'ko': """안녕하십니까 고객님, 저희 KKday를 이용해 주셔서 감사합니다.😇

고객님의 예약이 확인 되었으나,해당 날짜에 최소 정원 미달로 부득이하게 상품 제공이 어렵게 되었음을 안내드립니다.😥
현재 하기의 날짜 및 시간 예약이 가능한 점 참고 부탁드립니다.

👉【  】

*본 상품은 실시간으로 예약이 진행되어, 간혹 조기 마감이 될 수 있으니 이 점 양해 부탁드립니다.


1차 후보 : _______________
2차 후보 : _______________
(자동으로 고객님 순서에 따라 예약되며, 추후 공지하지 않습니다)

시간 변경 또는 예약 취소를 희망하실 경우 회신 부탁드립니다.

현재 예약이 아직 완료되지 않았으니 【D/L 】 (GMT+8)까지 회신 부탁드립니다. 답변을 받지 못한 경우 예약은 취소드리는 점 양해부탁드립니다..

다른 문의사항이 있으시면 언제든지 KKday 고객센터로 문의 주시길 바랍니다. 😊
감사합니다. 즐거운 자유여행은 KKday와 함께하세요. ：）✈️""",
    'ja': """KKdayをご利用いただきありがとうございます。

恐れ入りますが、ご予約いただきました日程は不催行のため手配ができませんでした。
現在以下の日程にて空きがある状況でございます。
‧ 【 】
※現時点での情報であり、空席を保証するものではございません。

つきまして
A：上記いずれかの日程へ変更を希望（ご希望順をお知らせください）
B：都合がつかないのでキャンセル希望
どちらかをご選択の上、お忙しい中恐れ入りますが台湾時間【D/L  】(GMT+8)までにご返信いただきますようよろしくお願いいたします。""",
}

# ── 排隊機制：最多同時跑 2 個瀏覽器 ──────────────────────
MAX_CONCURRENT = 2
_semaphore = threading.Semaphore(MAX_CONCURRENT)
_queue_lock = threading.Lock()
_waiting_count = 0
_stop_event = threading.Event()

def load_history():
    if not os.path.exists(HISTORY_FILE):
        return []
    with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_record(record):
    history = load_history()
    history.insert(0, record)
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history[:500], f, ensure_ascii=False, indent=2)
    if SHEETS_WEBHOOK:
        try:
            http_requests.post(SHEETS_WEBHOOK, json=record, timeout=10, allow_redirects=True)
        except Exception:
            pass

def calc_deadlines(s):
    now = datetime.now()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    diff = (datetime.strptime(s, '%Y-%m-%d') - today).days
    fmt = lambda d, t: d.strftime('%Y/%m/%d') + ' ' + t
    tomorrow = today + timedelta(days=1)
    
    if diff <= 1 and now.hour >= 17:
        ticket_dl = fmt(tomorrow, '10:00')
        follow_dl = fmt(today, '21:00')
    elif diff <= 3:
        if now.hour >= 17:
            ticket_dl = fmt(tomorrow, '11:00')
            follow_dl = fmt(tomorrow, '10:00')
        else:
            ticket_dl = fmt(today, '18:00')
            follow_dl = fmt(today, '17:00')
    else:
        d3 = today + timedelta(days=3)
        ticket_dl = fmt(d3, '17:00')
        follow_dl = fmt(d3, '12:00')
    
    return ticket_dl, follow_dl, diff <= 3, diff

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'run.log')

async def run_flow(order_id, progress, username, password, follow_type='page', op_dates='', supplier_order_id='', wantan_type='mansatisfied'):
    def push(msg, s='info'):
        progress.append({'msg': msg, 'status': s})
        line = f"  [{s}] {msg}"
        print(line)
        with open(LOG_FILE, 'a', encoding='utf-8') as lf:
            lf.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {line}\n")
        if s != 'error' and _stop_event.is_set():
            raise StopRequested('已手動停止')

    ticket_id = ''
    resolved_order_id = order_id.strip().upper() if order_id else ''

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context()
        page = await ctx.new_page()

        try:
            # ── 登入 ──────────────────────────────────────────
            push('登入 be2...')
            await page.goto('https://be2.kkday.com/login', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1000)

            # 點 Log In 按鈕，等 popup
            if await page.locator("text=Log In").count() > 0:
                async with ctx.expect_page() as popup_info:
                    await page.click("text=Log In")
                popup = await popup_info.value
                await popup.wait_for_load_state('networkidle')
                await popup.wait_for_timeout(1000)

                # 確保語系為繁體中文(台灣)，避免外國同事預設語系不同導致後續點擊失敗
                lang_sel = popup.locator('select')
                if await lang_sel.count() > 0:
                    await lang_sel.select_option(label='繁體中文(台灣)')
                    await popup.wait_for_timeout(300)

                await popup.fill("input[type='email']", username)
                await popup.wait_for_timeout(500)
                await popup.fill("input[type='password']", password)
                await popup.wait_for_timeout(500)
                submit = popup.locator("button[type='submit'], button:has-text('登入'), button:has-text('Log In'), button:has-text('Sign In')")
                await submit.first.click()
                await popup.wait_for_load_state('networkidle')
                await page.wait_for_timeout(6000)
                await page.goto('https://be2.kkday.com/v3/crm/dashboard', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(2000)
                if 'login' in page.url:
                    raise Exception('登入失敗，請確認帳號密碼')
                push('登入成功', 'ok')
            else:
                push('已登入', 'ok')

            # ── 確認登入成功 ───────────────────────────────────
            current = page.url
            push(f'登入後頁面：{current}')
            if 'login' in current:
                raise Exception('登入失敗，請確認帳號密碼')

            # ── 供應商編號 → 查 KKday 訂單編號 ───────────────
            if supplier_order_id and not resolved_order_id:
                push(f'透過供應商編號 {supplier_order_id} 查詢訂單...')
                await page.goto('https://be2.kkday.com/order/order_list', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(1500)
                supplier_input = page.locator("input[placeholder*='供應商訂單編號']")
                if await supplier_input.count() == 0:
                    supplier_input = page.locator("input[id*='supplier'], input[name*='supplier']").first
                if await supplier_input.count() > 0:
                    await supplier_input.fill(supplier_order_id)
                    await page.wait_for_timeout(300)
                else:
                    raise Exception('找不到供應商訂單編號欄位')
                await page.click("button:has-text('查詢')")
                await page.wait_for_timeout(2500)
                rows = page.locator("table tbody tr")
                row_count = await rows.count()
                if row_count > 0:
                    found_ids = []
                    for i in range(row_count):
                        row_text = await rows.nth(i).inner_text()
                        for m in re.findall(r'\d{2}KK\w+', row_text):
                            if m not in found_ids:
                                found_ids.append(m)
                    if len(found_ids) == 1:
                        resolved_order_id = found_ids[0]
                        push(f'找到訂單編號：{resolved_order_id}', 'ok')
                    elif len(found_ids) > 1:
                        raise Exception(f'供應商編號 {supplier_order_id} 對應多筆訂單（{", ".join(found_ids)}），請直接填入 KKday 訂單編號欄位')
                    else:
                        raise Exception(f'供應商編號 {supplier_order_id} 找不到對應的 KKday 訂單編號，請改填入 KKday 訂單編號欄位')
                else:
                    raise Exception(f'供應商編號 {supplier_order_id} 查無訂單，請改填入 KKday 訂單編號欄位')
            if not resolved_order_id:
                raise Exception('請輸入 KKday 訂單編號或供應商編號')

            use_date = ''
            urgent = False
            diff = 0

            # ── 前往客服操作台，新增工單 ───────────────────────
            push('前往客服操作台...')
            await page.goto('https://be2.kkday.com/v3/crm/dashboard?taskStatus=3', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1500)

            push('點擊新增工單...')
            await page.click("button:has-text('新增工單')")
            await page.wait_for_timeout(2000)
            await page.screenshot(path='modal_debug.png')

            # 選「有訂單」radio button
            await page.locator("label:has-text('有訂單')").click()
            await page.wait_for_timeout(1000)

            # 填訂單編號
            order_input = page.locator("input[placeholder='Please enter text']").first
            await order_input.fill(resolved_order_id)
            await page.keyboard.press('Tab')
            await page.wait_for_timeout(1000)

            # 選工單分類（全部用 JS，避免 get_by_text 30s timeout）
            if wantan_type == 'mituan':
                push('選工單分類：訂單異動→供應商通知→改期...')
            else:
                push('選工單分類：訂單異動→額滿→挽單...')
            await page.evaluate("""() => {
                const labels = Array.from(document.querySelectorAll('label, span, div'));
                const lbl = labels.find(e => e.textContent.includes('工單分類') && e.offsetParent !== null);
                if (lbl) {
                    let el = lbl;
                    for (let i = 0; i < 6; i++) {
                        el = el.parentElement;
                        if (!el) break;
                        const inp = el.querySelector('input.k-cascader__search-input');
                        if (inp && inp.offsetParent !== null) { inp.click(); return; }
                    }
                }
            }""")
            await page.wait_for_timeout(800)
            # L1：訂單異動
            found_l1 = await page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('li, div, span'));
                const el = els.find(e => e.textContent.trim() === '訂單異動' && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""")
            if not found_l1:
                raise Exception('找不到「訂單異動」分類選項')
            push('訂單異動 ✓')
            await page.wait_for_timeout(800)

            if wantan_type == 'mituan':
                # L2：供應商通知
                found_l2 = await page.evaluate("""() => {
                    const els = Array.from(document.querySelectorAll('li, div, span'));
                    const el = els.find(e => e.textContent.trim() === '供應商通知' && e.offsetParent !== null);
                    if (el) { el.click(); return true; }
                    return false;
                }""")
                if not found_l2:
                    raise Exception('找不到「供應商通知」分類選項')
                push('供應商通知 ✓')
                await page.wait_for_timeout(400)
                # L3：改期
                found_l3 = await page.evaluate("""() => {
                    const els = Array.from(document.querySelectorAll('div.text-ellipsis, li, span'));
                    const el = els.find(e => e.textContent.trim() === '改期' && e.offsetParent !== null);
                    if (el) { el.click(); return true; }
                    return false;
                }""")
                if not found_l3:
                    raise Exception('找不到「改期」選項，請確認工單分類下拉是否正確展開')
                push('改期 ✓')
            else:
                # L2：額滿 / 售罄
                found_l2 = await page.evaluate("""() => {
                    const els = Array.from(document.querySelectorAll('li'));
                    const el = els.find(e => e.textContent.trim().includes('額滿') && e.offsetParent !== null);
                    if (el) { el.click(); return true; }
                    return false;
                }""")
                if not found_l2:
                    raise Exception('找不到「額滿 / 售罄」分類選項')
                push('額滿 ✓')
                await page.wait_for_timeout(800)
                # L3：挽單
                found_l3 = await page.evaluate("""() => {
                    const els = Array.from(document.querySelectorAll('li, div, span'));
                    const el = els.find(e => e.textContent.trim() === '挽單' && e.offsetParent !== null);
                    if (el) { el.click(); return true; }
                    return false;
                }""")
                if not found_l3:
                    raise Exception('找不到「挽單」選項，請確認工單分類下拉是否正確展開')
                push('挽單 ✓')
            await page.wait_for_timeout(1500)

            # 最晚處理時間 - 點燈泡自動帶入
            push('點擊最晚處理時間燈泡...')
            await page.locator("button[class*='k-btn--orange']").first.click()
            await page.wait_for_timeout(3000)
            tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                push('燈泡未帶入，重試一次...')
                await page.locator("button[class*='k-btn--orange']").first.click()
                await page.wait_for_timeout(3000)
                tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                raise Exception('最晚處理時間燈泡點擊後仍為空，請確認工單分類與商品類型已選擇')
            push(f'最晚處理時間：{tdl}')

            await page.screenshot(path="before_confirm.png")
            push('確認建立工單...')
            # 用座標點確認按鈕
            await page.mouse.click(889, 651)
            await page.wait_for_timeout(2000)
            # 如果還有確認彈窗
            confirm2 = page.locator("button:has-text('確認')")
            if await confirm2.count() > 0:
                await confirm2.first.click()
                await page.wait_for_timeout(2000)
            push('工單建立成功！', 'ok')

            # ── 找工單 ID ──────────────────────────────────────
            await page.wait_for_timeout(4000)
            ticket_id = ''
            rows = page.locator("table tbody tr")
            invalid_vals = {'No Data', 'no data', 'NO DATA', '-', ''}
            for i in range(await rows.count()):
                if resolved_order_id in await rows.nth(i).inner_text():
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals:
                        ticket_id = val
                    break
            if not ticket_id:
                for i in range(await rows.count()):
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals and val and val[0].isdigit():
                        ticket_id = val
                        break
            if not ticket_id:
                ticket_id = '（請手動查看）'
            push(f'工單 ID：{ticket_id}', 'ok')
            if '請手動' in ticket_id:
                raise Exception('無法取得工單 ID，請手動至 BE2 查看並繼續後續步驟')

            push('點進工單 → 開始處理...')
            # 先按 Escape 關掉所有 modal/datepicker
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
            # 點工單 ID 讓右側面板顯示
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)
            # 點開始處理
            await page.click("button:has-text('開始處理')")
            await page.wait_for_timeout(800)
            # 第一個確認
            await page.locator("button:has-text('確認')").first.click()
            await page.wait_for_timeout(1000)
            # 第二個確認（操作成功彈窗）
            if await page.locator("button:has-text('確認')").count() > 0:
                await page.locator("button:has-text('確認')").first.click()
                await page.wait_for_timeout(1000)
            push('狀態更新為處理中', 'ok')
            await page.wait_for_timeout(2000)

            # ── 後拋 ───────────────────────────────────────────
            push('點擊後拋...')
            # 點「處理中」tab
            await page.locator("text=處理中").first.click()
            await page.wait_for_timeout(1000)
            await page.screenshot(path="processing_tab.png")

            # 點工單 ID 讓右側面板出現
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)
            await page.screenshot(path="ticket_panel.png")

            # 確認後拋按鈕存在再點（精確匹配，避免誤點「後拋中」tab）
            await page.wait_for_selector("button:has-text('後拋')", timeout=8000)
            push(f'後拋前頁面：{page.url}')
            await page.get_by_role("button", name="後拋", exact=True).click()
            await page.wait_for_timeout(2000)
            await page.screenshot(path="follow_debug.png")
            push(f'後拋後頁面：{page.url}')

            # 確認 modal 出現
            modal_ok = await page.locator("text=後拋處理人").count() > 0
            push(f'後拋 modal 出現：{modal_ok}')
            if not modal_ok:
                raise Exception(f'後拋 modal 未出現，當前頁面：{page.url}，請查看 follow_debug.png')

                        # 選自動分派
            auto = page.locator("label:has-text('自動分派')")
            if await auto.count() > 0:
                await auto.click()
            await page.wait_for_timeout(300)

            # 後拋內容
            post_ta = page.locator('textarea').first
            if await post_ta.count() > 0:
                await post_ta.fill('未成團挽單' if wantan_type == 'mituan' else '額滿挽單')
            await page.wait_for_timeout(300)

            # 後拋截止時間 - 點燈泡自動帶入
            push('點擊後拋截止時間燈泡...')
            await page.locator("button[class*='k-btn--orange']").first.click()
            await page.wait_for_timeout(2000)
            fdl = await page.locator("input[placeholder='Select date']").last.input_value()
            if not fdl:
                raise Exception('後拋截止時間燈泡點擊後仍為空')
            push(f'後拋截止時間：{fdl}')

            # 此後拋需要通知旅客？→ 選「是」
            await page.evaluate("""() => {
                const labels = Array.from(document.querySelectorAll('label'));
                const yesLabel = labels.find(l => l.textContent.trim() === '是');
                if (yesLabel) yesLabel.click();
            }""")
            await page.wait_for_timeout(300)

            # 誰來發送對客通知 → 系統自動發送
            sys_send = page.locator("label:has-text('系統自動發送')")
            if await sys_send.count() > 0:
                await sys_send.click()
            await page.wait_for_timeout(500)

            # 讀取旅客語系（支援 zh-tw / zh-cn / CN / zh-hk / en / ko / ja）
            customer_lang = await page.evaluate("""() => {
                const allEls = Array.from(document.querySelectorAll('td, span, div'));
                for (const el of allEls) {
                    const t = el.textContent.trim();
                    if (/^(zh-tw|zh-cn|zh-hk|en|ko|ja|CN)$/i.test(t)) return t.toLowerCase();
                }
                return 'en';
            }""")
            # 統一 CN → zh-cn
            if customer_lang == 'cn':
                customer_lang = 'zh-cn'
            push(f'旅客語系：{customer_lang}')

            # 計算通知 D/L（後拋截止 - 2小時）
            fdl_dt = datetime.strptime(fdl, '%Y/%m/%d %H:%M')
            notif_dl = (fdl_dt - timedelta(hours=2)).strftime('%Y/%m/%d %H:%M')

            # 取通知內容，替換 D/L（各語系空白數量不同，用 regex）與 OP 場次
            push(f'挽單類型：{"未成團" if wantan_type == "mituan" else "額滿"}，可改類型：{follow_type}')
            if wantan_type == 'mituan':
                notif_tmpl = NOTIF_TEMPLATES_MITUAN
                notif_content = notif_tmpl.get(customer_lang, notif_tmpl.get('en', ''))
                notif_subject = NOTIF_SUBJECT_MITUAN
            else:
                notif_tmpl = NOTIF_TEMPLATES.get(follow_type, NOTIF_TEMPLATES['page'])
                notif_content = notif_tmpl.get(customer_lang, notif_tmpl.get('en', ''))
                notif_subject = NOTIF_SUBJECT
            notif_content = re.sub(r'【\s*D/L\s*】', f'【{notif_dl}】', notif_content)
            if op_dates:
                notif_content = re.sub(r'👉【\s*】', f'👉【{op_dates}】', notif_content)
                notif_content = re.sub(r'‧ 【\s*】', f'‧ 【{op_dates}】', notif_content)

            # 填通知主旨
            await page.evaluate("""(subject) => {
                const labels = Array.from(document.querySelectorAll('label'));
                const lb = labels.find(l => l.textContent.trim().includes('通知主旨'));
                let input = lb ? (lb.parentElement.querySelector('input') ||
                                  lb.nextElementSibling?.querySelector('input')) : null;
                if (!input) input = document.querySelector('input[placeholder*="主旨"]');
                if (input) {
                    input.focus(); input.value = subject;
                    input.dispatchEvent(new Event('input', {bubbles: true}));
                    input.dispatchEvent(new Event('change', {bubbles: true}));
                }
            }""", notif_subject)
            await page.wait_for_timeout(300)
            push('通知主旨已填入')

            # 填通知內容
            await page.evaluate("""(content) => {
                const labels = Array.from(document.querySelectorAll('label'));
                const lb = labels.find(l => l.textContent.trim().includes('通知內容'));
                let ta = lb ? (lb.parentElement.querySelector('textarea') ||
                               lb.nextElementSibling?.querySelector('textarea')) : null;
                if (!ta) {
                    const tas = Array.from(document.querySelectorAll('textarea'));
                    ta = tas[tas.length - 1];
                }
                if (ta) {
                    ta.focus(); ta.value = content;
                    ta.dispatchEvent(new Event('input', {bubbles: true}));
                    ta.dispatchEvent(new Event('change', {bubbles: true}));
                }
            }""", notif_content)
            await page.wait_for_timeout(300)
            push(f'通知內容已填入（語系：{customer_lang}，D/L：{notif_dl}）')

            # 確認後拋
            await page.get_by_role("button", name="確認").last.click()
            await page.wait_for_timeout(1500)

            # ── 訂單備註 ───────────────────────────────────────
            push('填入訂單備註...')
            note = f'挽單自動化\n#{ticket_id}\n最晚處理\n{tdl}\n後拋截止\n{fdl}'
            await page.goto(
                f'https://be2.kkday.com/order/order_view/{resolved_order_id}',
                wait_until='networkidle', timeout=30000
            )
            await page.wait_for_timeout(2000)
            await page.locator("a[role='tab']:has-text('訂單備註')").first.click()
            await page.wait_for_timeout(2000)
            # 用 JS 找第一個可見的 textarea 填入備註
            filled = await page.evaluate("""(note) => {
                const tas = Array.from(document.querySelectorAll('textarea'));
                const visible = tas.find(ta => ta.offsetParent !== null && ta.offsetHeight > 0);
                if (!visible) return false;
                visible.focus();
                visible.value = note;
                visible.dispatchEvent(new Event('input', {bubbles: true}));
                visible.dispatchEvent(new Event('change', {bubbles: true}));
                return true;
            }""", note)
            if not filled:
                raise Exception('找不到訂單備註 textarea，請確認頁面已載入')
            await page.wait_for_timeout(500)
            await page.locator("button:has-text('記錄備註')").click()
            await page.wait_for_timeout(1500)
            push('訂單備註已記錄', 'ok')

            push(f'全部完成！工單 ID：{ticket_id}', 'ok')
            await browser.close()
            return {'success': True, 'ticket_id': ticket_id, 'use_date': use_date,
                    'ticket_deadline': tdl, 'follow_deadline': fdl,
                    'is_urgent': urgent, 'diff_days': diff, 'order_id': resolved_order_id}

        except StopRequested:
            push('已手動停止', 'error')
            try:
                await browser.close()
            except Exception:
                pass
            return {'success': False, 'error': '已手動停止', 'ticket_id': ticket_id, 'order_id': resolved_order_id}

        except Exception as e:
            push(f'錯誤：{str(e)}', 'error')
            await browser.close()
            return {'success': False, 'error': str(e), 'ticket_id': ticket_id, 'order_id': resolved_order_id}


async def run_notification_flow(order_id, supplier_order_id, notification_content, progress, username, password):
    def push(msg, s='info'):
        progress.append({'msg': msg, 'status': s})
        line = f"  [{s}] {msg}"
        print(line)
        with open(LOG_FILE, 'a', encoding='utf-8') as lf:
            lf.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {line}\n")
        if s != 'error' and _stop_event.is_set():
            raise StopRequested('已手動停止')

    ticket_id = ''
    resolved_order_id = order_id.strip().upper() if order_id else ''

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context()
        page = await ctx.new_page()

        try:
            # ── 登入 ──────────────────────────────────────────
            push('登入 be2...')
            await page.goto('https://be2.kkday.com/login', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1000)
            if await page.locator("text=Log In").count() > 0:
                async with ctx.expect_page() as popup_info:
                    await page.click("text=Log In")
                popup = await popup_info.value
                await popup.wait_for_load_state('networkidle')
                await popup.wait_for_timeout(1000)

                # 確保語系為繁體中文(台灣)
                lang_sel = popup.locator('select')
                if await lang_sel.count() > 0:
                    await lang_sel.select_option(label='繁體中文(台灣)')
                    await popup.wait_for_timeout(300)

                await popup.fill("input[type=\'email\']", username)
                await popup.wait_for_timeout(500)
                await popup.fill("input[type=\'password\']", password)
                await popup.wait_for_timeout(500)
                submit = popup.locator("button[type='submit'], button:has-text('登入'), button:has-text('Log In'), button:has-text('Sign In')")
                await submit.first.click()
                await popup.wait_for_load_state('networkidle')
                await page.wait_for_timeout(6000)
                await page.goto('https://be2.kkday.com/v3/crm/dashboard', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(2000)
                if 'login' in page.url:
                    raise Exception('登入失敗，請確認帳號密碼')
                push('登入成功', 'ok')
            else:
                push('已登入', 'ok')

            # ── 供應商編號 → 查 KKday 訂單編號 ───────────────
            if supplier_order_id and not resolved_order_id:
                push(f'透過供應商編號 {supplier_order_id} 查詢訂單...')
                await page.goto('https://be2.kkday.com/order/order_list', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(1500)
                supplier_input = page.locator("input[placeholder*=\'供應商訂單編號\']")
                if await supplier_input.count() == 0:
                    supplier_input = page.locator("input[id*=\'supplier\'], input[name*=\'supplier\']").first
                if await supplier_input.count() > 0:
                    await supplier_input.fill(supplier_order_id)
                    await page.wait_for_timeout(300)
                else:
                    raise Exception('找不到供應商訂單編號欄位')
                await page.click("button:has-text(\'查詢\')")
                await page.wait_for_timeout(2500)
                rows = page.locator("table tbody tr")
                row_count = await rows.count()
                if row_count > 0:
                    found_ids = []
                    for i in range(row_count):
                        row_text = await rows.nth(i).inner_text()
                        for m in re.findall(r'\d{2}KK\w+', row_text):
                            if m not in found_ids:
                                found_ids.append(m)
                    if len(found_ids) == 1:
                        resolved_order_id = found_ids[0]
                        push(f'找到訂單編號：{resolved_order_id}', 'ok')
                    elif len(found_ids) > 1:
                        raise Exception(f'供應商編號 {supplier_order_id} 對應多筆訂單（{", ".join(found_ids)}），請直接填入 KKday 訂單編號欄位')
                    else:
                        raise Exception(f'供應商編號 {supplier_order_id} 找不到對應的 KKday 訂單編號，請改填入 KKday 訂單編號欄位')
                else:
                    raise Exception(f'供應商編號 {supplier_order_id} 查無訂單，請改填入 KKday 訂單編號欄位')
            if not resolved_order_id:
                raise Exception('請輸入 KKday 訂單編號或供應商編號')

            use_date = ''
            urgent = False
            diff = 0

            # ── 前往客服操作台，新增工單 ───────────────────────
            push('前往客服操作台...')
            await page.goto('https://be2.kkday.com/v3/crm/dashboard?taskStatus=3', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1500)
            push('點擊新增工單...')
            await page.click("button:has-text(\'新增工單\')")
            await page.wait_for_timeout(2000)
            await page.locator("label:has-text(\'有訂單\')").click()
            await page.wait_for_timeout(1000)
            order_input = page.locator("input[placeholder=\'Please enter text\']").first
            await order_input.fill(resolved_order_id)
            await page.keyboard.press('Tab')
            await page.wait_for_timeout(1000)

            # 選工單分類：供應商自理訊息 → 供應商通知 → 轉達行前注意事項
            push('選工單分類：供應商自理訊息→供應商通知→轉達行前注意事項...')
            await page.evaluate("""() => {
                const labels = Array.from(document.querySelectorAll('label, span, div'));
                const lbl = labels.find(e => e.textContent.includes('工單分類') && e.offsetParent !== null);
                if (lbl) {
                    let el = lbl;
                    for (let i = 0; i < 6; i++) {
                        el = el.parentElement;
                        if (!el) break;
                        const inp = el.querySelector('input.k-cascader__search-input');
                        if (inp && inp.offsetParent !== null) { inp.click(); return; }
                    }
                }
            }""")
            await page.wait_for_timeout(800)
            found_l1 = await page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('li, div, span'));
                const el = els.find(e => e.textContent.trim() === '供應商自理訊息' && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""")
            if not found_l1:
                raise Exception('找不到「供應商自理訊息」分類選項，請確認下拉選單已展開')
            push('供應商自理訊息 ✓')
            await page.wait_for_timeout(400)
            found_l2 = await page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('li, div, span'));
                const el = els.find(e => e.textContent.trim() === '供應商通知' && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""")
            if not found_l2:
                raise Exception('找不到「供應商通知」分類選項，請確認下拉選單已展開')
            push('供應商通知 ✓')
            await page.wait_for_timeout(600)
            found_l3 = await page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('div.text-ellipsis, li, span'));
                const el = els.find(e => e.textContent.trim() === '轉達行前注意事項' && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""")
            if not found_l3:
                raise Exception('找不到「轉達行前注意事項」選項，請確認工單分類下拉是否正確展開')
            push('轉達行前注意事項 ✓')
            await page.wait_for_timeout(1500)

            # 最晚處理時間 - 點燈泡自動帶入
            push('點擊最晚處理時間燈泡...')
            await page.locator("button[class*='k-btn--orange']").first.click()
            await page.wait_for_timeout(3000)
            tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                push('燈泡未帶入，重試一次...')
                await page.locator("button[class*='k-btn--orange']").first.click()
                await page.wait_for_timeout(3000)
                tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                raise Exception('最晚處理時間燈泡點擊後仍為空，請確認工單分類與商品類型已選擇')
            push(f'最晚處理時間：{tdl}', 'ok')

            push('確認建立工單...')
            await page.mouse.click(889, 651)
            await page.wait_for_timeout(2000)
            confirm2 = page.locator("button:has-text(\'確認\')")
            if await confirm2.count() > 0:
                await confirm2.first.click()
                await page.wait_for_timeout(2000)
            push('工單建立成功！', 'ok')

            # ── 找工單 ID ──────────────────────────────────────
            await page.wait_for_timeout(4000)
            ticket_id = ''
            rows = page.locator("table tbody tr")
            invalid_vals = {'No Data', 'no data', 'NO DATA', '-', ''}
            for i in range(await rows.count()):
                if resolved_order_id in await rows.nth(i).inner_text():
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals:
                        ticket_id = val
                    break
            if not ticket_id:
                for i in range(await rows.count()):
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals and val and val[0].isdigit():
                        ticket_id = val
                        break
            if not ticket_id:
                ticket_id = '（請手動查看）'
            push(f'工單 ID：{ticket_id}', 'ok')
            if '請手動' in ticket_id:
                raise Exception('無法取得工單 ID，請手動至 BE2 查看並繼續後續步驟')

            # ── 開始處理 ───────────────────────────────────────
            push('點進工單 → 開始處理...')
            await page.goto('https://be2.kkday.com/v3/crm/dashboard?taskStatus=3', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1500)
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)
            await page.click("button:has-text(\'開始處理\')")
            await page.wait_for_timeout(800)
            await page.locator("button:has-text(\'確認\')").first.click()
            await page.wait_for_timeout(1000)
            if await page.locator("button:has-text(\'確認\')").count() > 0:
                await page.locator("button:has-text(\'確認\')").first.click()
                await page.wait_for_timeout(1000)
            push('狀態更新為處理中', 'ok')
            await page.wait_for_timeout(1500)
            # 關閉「操作成功」彈窗
            if await page.locator("button:has-text('確認')").count() > 0:
                await page.locator("button:has-text('確認')").first.click()
                await page.wait_for_timeout(800)

            # ── 對客通知 ───────────────────────────────────────
            push('點擊對客通知...')
            await page.get_by_role("button", name="對客通知").click()
            await page.wait_for_timeout(2000)

            # 收件人已自動帶入，直接下一步
            await page.get_by_role("button", name="下一步").click()
            await page.wait_for_timeout(1500)

            # 填通知主旨
            notif_subject = '對客通知 Customer Notification'
            await page.evaluate("""(subject) => {
                const labels = Array.from(document.querySelectorAll('label'));
                const lb = labels.find(l => l.textContent.trim().includes('通知主旨'));
                let input = lb ? (lb.parentElement.querySelector('input') ||
                                  lb.nextElementSibling?.querySelector('input')) : null;
                if (!input) input = document.querySelector('input[placeholder*="主旨"]');
                if (input) {
                    input.focus(); input.value = subject;
                    input.dispatchEvent(new Event('input', {bubbles: true}));
                    input.dispatchEvent(new Event('change', {bubbles: true}));
                }
            }""", notif_subject)
            await page.wait_for_timeout(300)
            push('通知主旨已填入', 'ok')

            # 填通知內容（使用者輸入的行前通知內容）
            await page.evaluate("""(content) => {
                const labels = Array.from(document.querySelectorAll('label'));
                const lb = labels.find(l => l.textContent.trim().includes('通知內容'));
                let ta = lb ? (lb.parentElement.querySelector('textarea') ||
                               lb.nextElementSibling?.querySelector('textarea')) : null;
                if (!ta) {
                    const tas = Array.from(document.querySelectorAll('textarea'));
                    ta = tas[tas.length - 1];
                }
                if (ta) {
                    ta.focus(); ta.value = content;
                    ta.dispatchEvent(new Event('input', {bubbles: true}));
                    ta.dispatchEvent(new Event('change', {bubbles: true}));
                }
            }""", notification_content)
            await page.wait_for_timeout(300)
            push('通知內容已填入', 'ok')

            # 下一步 → 預覽&發送
            await page.get_by_role("button", name="下一步").click()
            await page.wait_for_timeout(1500)

            # 確認發送
            await page.get_by_role("button", name="確認發送").click()
            await page.wait_for_timeout(2000)
            push('對客通知已發送', 'ok')

            # ── 關閉工單 ───────────────────────────────────────
            push('關閉工單...')
            await page.goto('https://be2.kkday.com/v3/crm/dashboard?taskStatus=3', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1500)
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)
            await page.get_by_role("button", name="關閉", exact=True).click()
            await page.wait_for_timeout(1000)
            if await page.locator("button:has-text('確認')").count() > 0:
                await page.locator("button:has-text('確認')").first.click()
                await page.wait_for_timeout(1000)
            push('工單已關閉', 'ok')

            # ── 訂單備註 ───────────────────────────────────────
            push('填入訂單備註...')
            note = f'對客通知自動化\n#{ticket_id}\n最晚處理\n{tdl}'
            await page.goto(f'https://be2.kkday.com/order/order_view/{resolved_order_id}', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(2000)
            await page.locator("a[role='tab']:has-text('訂單備註')").first.click()
            await page.wait_for_timeout(2000)
            filled = await page.evaluate("""(note) => {
                const tas = Array.from(document.querySelectorAll('textarea'));
                const visible = tas.find(ta => ta.offsetParent !== null && ta.offsetHeight > 0);
                if (!visible) return false;
                visible.focus();
                visible.value = note;
                visible.dispatchEvent(new Event('input', {bubbles: true}));
                visible.dispatchEvent(new Event('change', {bubbles: true}));
                return true;
            }""", note)
            if not filled:
                raise Exception('找不到訂單備註 textarea')
            await page.wait_for_timeout(500)
            await page.locator("button:has-text(\'記錄備註\')").click()
            await page.wait_for_timeout(1500)
            push('訂單備註已記錄', 'ok')

            push(f'全部完成！工單 ID：{ticket_id}', 'ok')
            await browser.close()
            return {'success': True, 'ticket_id': ticket_id, 'use_date': use_date,
                    'ticket_deadline': tdl,
                    'is_urgent': urgent, 'diff_days': diff, 'order_id': resolved_order_id}

        except StopRequested:
            push('已手動停止', 'error')
            try:
                await browser.close()
            except Exception:
                pass
            return {'success': False, 'error': '已手動停止', 'ticket_id': ticket_id, 'order_id': resolved_order_id}

        except Exception as e:
            push(f'錯誤：{str(e)}', 'error')
            await browser.close()
            return {'success': False, 'error': str(e), 'ticket_id': ticket_id, 'order_id': resolved_order_id}


async def run_general_single(order_id, supplier_order_id, cat_l1, cat_l2, cat_l3,
                              notif_content, progress, username, password, operator='', notify_customer='yes'):
    def push(msg, s='info'):
        progress.append({'msg': msg, 'status': s})
        line = f"  [{s}] {msg}"
        print(line)
        with open(LOG_FILE, 'a', encoding='utf-8') as lf:
            lf.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {line}\n")
        if s != 'error' and _stop_event.is_set():
            raise StopRequested('已手動停止')

    ticket_id = ''
    resolved_order_id = order_id.strip().upper() if order_id else ''

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        ctx = await browser.new_context()
        page = await ctx.new_page()

        try:
            # ── 登入 ──────────────────────────────────────────
            push('登入 be2...')
            await page.goto('https://be2.kkday.com/login', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1000)

            if await page.locator("text=Log In").count() > 0:
                async with ctx.expect_page() as popup_info:
                    await page.click("text=Log In")
                popup = await popup_info.value
                await popup.wait_for_load_state('networkidle')
                await popup.wait_for_timeout(1000)

                # 確保語系為繁體中文(台灣)
                lang_sel = popup.locator('select')
                if await lang_sel.count() > 0:
                    await lang_sel.select_option(label='繁體中文(台灣)')
                    await popup.wait_for_timeout(300)

                await popup.fill("input[type='email']", username)
                await popup.wait_for_timeout(500)
                await popup.fill("input[type='password']", password)
                await popup.wait_for_timeout(500)
                submit = popup.locator("button[type='submit'], button:has-text('登入'), button:has-text('Log In'), button:has-text('Sign In')")
                await submit.first.click()
                await popup.wait_for_load_state('networkidle')
                await page.wait_for_timeout(6000)
                await page.goto('https://be2.kkday.com/v3/crm/dashboard', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(2000)
                if 'login' in page.url:
                    raise Exception('登入失敗，請確認帳號密碼')
                push('登入成功', 'ok')
            else:
                push('已登入', 'ok')

            current = page.url
            if 'login' in current:
                raise Exception('登入失敗，請確認帳號密碼')

            # ── 供應商編號 → 查 KKday 訂單編號 ───────────────
            if supplier_order_id and not resolved_order_id:
                push(f'透過供應商編號 {supplier_order_id} 查詢訂單...')
                await page.goto('https://be2.kkday.com/order/order_list', wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(1500)
                supplier_input = page.locator("input[placeholder*='供應商訂單編號']")
                if await supplier_input.count() == 0:
                    supplier_input = page.locator("input[id*='supplier'], input[name*='supplier']").first
                if await supplier_input.count() > 0:
                    await supplier_input.fill(supplier_order_id)
                    await page.wait_for_timeout(300)
                else:
                    raise Exception('找不到供應商訂單編號欄位')
                await page.click("button:has-text('查詢')")
                await page.wait_for_timeout(2500)
                rows = page.locator("table tbody tr")
                row_count = await rows.count()
                if row_count > 0:
                    found_ids = []
                    for i in range(row_count):
                        row_text = await rows.nth(i).inner_text()
                        for m in re.findall(r'\d{2}KK\w+', row_text):
                            if m not in found_ids:
                                found_ids.append(m)
                    if len(found_ids) == 1:
                        resolved_order_id = found_ids[0]
                        push(f'找到訂單編號：{resolved_order_id}', 'ok')
                    elif len(found_ids) > 1:
                        raise Exception(f'供應商編號 {supplier_order_id} 對應多筆訂單（{", ".join(found_ids)}），請直接填入 KKday 訂單編號欄位')
                    else:
                        raise Exception(f'供應商編號 {supplier_order_id} 找不到對應的 KKday 訂單編號，請改填入 KKday 訂單編號欄位')
                else:
                    raise Exception(f'供應商編號 {supplier_order_id} 查無訂單，請改填入 KKday 訂單編號欄位')
            if not resolved_order_id:
                raise Exception('請輸入 KKday 訂單編號或供應商編號')

            use_date = ''
            urgent = False
            diff = 0

            # ── 前往客服操作台，新增工單 ───────────────────────
            push('前往客服操作台...')
            await page.goto('https://be2.kkday.com/v3/crm/dashboard?taskStatus=3', wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1500)

            push('點擊新增工單...')
            await page.click("button:has-text('新增工單')")
            await page.wait_for_timeout(2000)

            await page.locator("label:has-text('有訂單')").click()
            await page.wait_for_timeout(1000)

            order_input = page.locator("input[placeholder='Please enter text']").first
            await order_input.fill(resolved_order_id)
            await page.keyboard.press('Tab')
            await page.wait_for_timeout(1000)

            # ── 選工單分類（全部用 JS，避免 get_by_text 30s timeout）─────
            push(f'選工單分類：{cat_l1}→{cat_l2}→{cat_l3}...')

            await page.evaluate("""() => {
                const labels = Array.from(document.querySelectorAll('label, span, div'));
                const lbl = labels.find(e => e.textContent.includes('工單分類') && e.offsetParent !== null);
                if (lbl) {
                    let el = lbl;
                    for (let i = 0; i < 6; i++) {
                        el = el.parentElement;
                        if (!el) break;
                        const inp = el.querySelector('input.k-cascader__search-input');
                        if (inp && inp.offsetParent !== null) { inp.click(); return; }
                    }
                }
            }""")
            await page.wait_for_timeout(1000)
            found_l1 = await page.evaluate("""(l1) => {
                const els = Array.from(document.querySelectorAll('li, div, span'));
                const el = els.find(e => e.textContent.trim() === l1 && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""", cat_l1)
            if not found_l1:
                raise Exception(f'找不到「{cat_l1}」分類選項，請確認下拉選單已展開')
            push(f'{cat_l1} ✓')
            await page.wait_for_timeout(500)
            found_l2 = await page.evaluate("""(l2) => {
                const els = Array.from(document.querySelectorAll('li, div, span'));
                const el = els.find(e => e.textContent.trim() === l2 && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""", cat_l2)
            if not found_l2:
                raise Exception(f'找不到「{cat_l2}」分類選項，請確認 {cat_l1} 已選擇')
            push(f'{cat_l2} ✓')
            await page.wait_for_timeout(600)
            found_l3 = await page.evaluate("""(l3) => {
                const els = Array.from(document.querySelectorAll('div.text-ellipsis, li, span'));
                const el = els.find(e => e.textContent.trim() === l3 && e.offsetParent !== null);
                if (el) { el.click(); return true; }
                return false;
            }""", cat_l3)
            if not found_l3:
                raise Exception(f'找不到「{cat_l3}」選項，請確認工單分類下拉是否正確展開')
            await page.wait_for_timeout(2500)
            push(f'工單分類已選：{cat_l1}→{cat_l2}→{cat_l3}', 'ok')

            # ── 最晚處理時間燈泡 ───────────────────────────────
            push('點擊最晚處理時間燈泡...')

            async def click_deadline_bulb():
                clicked = await page.evaluate("""() => {
                    const labels = Array.from(document.querySelectorAll('label, span, div'));
                    const lbl = labels.find(l =>
                        l.textContent.trim().includes('最晚處理時間') && l.offsetParent !== null);
                    if (lbl) {
                        const container = lbl.closest('tr, [class*="form-item"], [class*="field"], [class*="row"]')
                                          || lbl.parentElement?.parentElement;
                        if (container) {
                            const btn = container.querySelector('button[class*="k-btn--orange"]');
                            if (btn) { btn.click(); return true; }
                        }
                    }
                    const allBtns = Array.from(document.querySelectorAll('button[class*="k-btn--orange"]'));
                    if (allBtns.length > 0) { allBtns[0].click(); return true; }
                    return false;
                }""")
                return clicked

            await click_deadline_bulb()
            await page.wait_for_timeout(3500)
            tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                push('燈泡未帶入，重試一次...')
                await click_deadline_bulb()
                await page.wait_for_timeout(4000)
                tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                push('燈泡未帶入，重試第二次...')
                await click_deadline_bulb()
                await page.wait_for_timeout(5000)
                tdl = await page.locator("input[placeholder='Select date']").first.input_value()
            if not tdl:
                raise Exception('最晚處理時間燈泡點擊後仍為空，請確認工單分類是否完整選擇')
            push(f'最晚處理時間：{tdl}', 'ok')

            push('確認建立工單...')
            confirmed = await page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button'));
                const btn = btns.find(b =>
                    b.textContent.trim() === '確認' &&
                    b.offsetParent !== null &&
                    !b.disabled
                );
                if (btn) { btn.click(); return true; }
                return false;
            }""")
            if not confirmed:
                await page.get_by_role("button", name="確認").first.click()
            await page.wait_for_timeout(2000)
            confirm2 = page.locator("button:has-text('確認')")
            if await confirm2.count() > 0:
                await confirm2.first.click()
                await page.wait_for_timeout(2000)
            push('工單建立成功！', 'ok')

            # ── 找工單 ID ──────────────────────────────────────
            await page.wait_for_timeout(4000)
            ticket_id = ''
            rows = page.locator("table tbody tr")
            invalid_vals = {'No Data', 'no data', 'NO DATA', '-', ''}
            for i in range(await rows.count()):
                if resolved_order_id in await rows.nth(i).inner_text():
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals:
                        ticket_id = val
                    break
            if not ticket_id:
                for i in range(await rows.count()):
                    val = (await rows.nth(i).locator('td').first.inner_text()).strip()
                    if val not in invalid_vals and val and val[0].isdigit():
                        ticket_id = val
                        break
            if not ticket_id:
                ticket_id = '（請手動查看）'
            push(f'工單 ID：{ticket_id}', 'ok')
            if '請手動' in ticket_id:
                raise Exception('無法取得工單 ID，請手動至 BE2 查看並繼續後續步驟')

            push('點進工單 → 開始處理...')
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)
            await page.click("button:has-text('開始處理')")
            await page.wait_for_timeout(800)
            await page.locator("button:has-text('確認')").first.click()
            await page.wait_for_timeout(1000)
            if await page.locator("button:has-text('確認')").count() > 0:
                await page.locator("button:has-text('確認')").first.click()
                await page.wait_for_timeout(1000)
            push('狀態更新為處理中', 'ok')
            await page.wait_for_timeout(2000)

            # ── 後拋 ───────────────────────────────────────────
            push('點擊後拋...')
            await page.locator("text=處理中").first.click()
            await page.wait_for_timeout(1000)
            await page.locator(f"text={ticket_id}").first.click()
            await page.wait_for_timeout(1500)

            await page.wait_for_selector("button:has-text('後拋')", timeout=8000)
            await page.get_by_role("button", name="後拋", exact=True).click()
            await page.wait_for_timeout(2000)

            modal_ok = await page.locator("text=後拋處理人").count() > 0
            if not modal_ok:
                raise Exception('後拋 modal 未出現')

            auto = page.locator("label:has-text('自動分派')")
            if await auto.count() > 0:
                await auto.click()
            await page.wait_for_timeout(300)

            # 後拋備註 — 填入 OP 輸入通知內容（兩條路徑相同）
            post_ta = page.locator('textarea').first
            if await post_ta.count() > 0:
                await post_ta.fill(notif_content)
            await page.wait_for_timeout(300)
            push('後拋備註已填入', 'ok')

            # 後拋截止時間燈泡
            push('點擊後拋截止時間燈泡...')
            await page.locator("button[class*='k-btn--orange']").first.click()
            await page.wait_for_timeout(2000)
            fdl = await page.locator("input[placeholder='Select date']").last.input_value()
            if not fdl:
                raise Exception('後拋截止時間燈泡點擊後仍為空')
            push(f'後拋截止時間：{fdl}', 'ok')

            if notify_customer == 'yes':
                # 此後拋需要通知旅客 → 是
                await page.evaluate("""() => {
                    const labels = Array.from(document.querySelectorAll('label'));
                    const yesLabel = labels.find(l => l.textContent.trim() === '是');
                    if (yesLabel) yesLabel.click();
                }""")
                await page.wait_for_timeout(300)
                push('此後拋需通知旅客 → 是')

                # 誰來發送對客通知 → 後拋處理人
                post_handler = page.locator("label:has-text('後拋處理人')")
                if await post_handler.count() > 0:
                    await post_handler.click()
                await page.wait_for_timeout(500)
                push('選擇後拋處理人', 'ok')

                # 通知主旨 — 自動帶入分類路徑
                notif_subject = f'{cat_l1} > {cat_l2} > {cat_l3}'
                await page.evaluate("""(subject) => {
                    const labels = Array.from(document.querySelectorAll('label'));
                    const lb = labels.find(l => l.textContent.trim().includes('通知主旨'));
                    let input = lb ? (lb.parentElement.querySelector('input') ||
                                      lb.nextElementSibling?.querySelector('input')) : null;
                    if (!input) input = document.querySelector('input[placeholder*="主旨"]');
                    if (input) {
                        input.focus(); input.value = subject;
                        input.dispatchEvent(new Event('input', {bubbles: true}));
                        input.dispatchEvent(new Event('change', {bubbles: true}));
                    }
                }""", notif_subject)
                await page.wait_for_timeout(300)
                push(f'通知主旨已填入：{notif_subject}')

                # 通知內容 — OP 提供
                await page.evaluate("""(content) => {
                    const labels = Array.from(document.querySelectorAll('label'));
                    const lb = labels.find(l => l.textContent.trim().includes('通知內容'));
                    let ta = lb ? (lb.parentElement.querySelector('textarea') ||
                                   lb.nextElementSibling?.querySelector('textarea')) : null;
                    if (!ta) {
                        const tas = Array.from(document.querySelectorAll('textarea'));
                        ta = tas[tas.length - 1];
                    }
                    if (ta) {
                        ta.focus(); ta.value = content;
                        ta.dispatchEvent(new Event('input', {bubbles: true}));
                        ta.dispatchEvent(new Event('change', {bubbles: true}));
                    }
                }""", notif_content)
                await page.wait_for_timeout(300)
                push('通知內容已填入')

            else:
                # 此後拋不需要通知旅客 → 否
                await page.evaluate("""() => {
                    const labels = Array.from(document.querySelectorAll('label'));
                    const noLabel = labels.find(l => l.textContent.trim() === '否');
                    if (noLabel) noLabel.click();
                }""")
                await page.wait_for_timeout(300)
                push('此後拋不需通知旅客 → 否')

            # 確認後拋
            await page.get_by_role("button", name="確認").last.click()
            await page.wait_for_timeout(1500)

            # ── 訂單備註 ───────────────────────────────────────
            push('填入訂單備註...')
            note = f'一般工單自動化\n#{ticket_id}\n分類：{cat_l1}>{cat_l2}>{cat_l3}\n最晚處理\n{tdl}\n後拋截止\n{fdl}'
            await page.goto(
                f'https://be2.kkday.com/order/order_view/{resolved_order_id}',
                wait_until='networkidle', timeout=30000
            )
            await page.wait_for_timeout(2000)
            await page.locator("a[role='tab']:has-text('訂單備註')").first.click()
            await page.wait_for_timeout(2000)
            filled = await page.evaluate("""(note) => {
                const tas = Array.from(document.querySelectorAll('textarea'));
                const visible = tas.find(ta => ta.offsetParent !== null && ta.offsetHeight > 0);
                if (!visible) return false;
                visible.focus();
                visible.value = note;
                visible.dispatchEvent(new Event('input', {bubbles: true}));
                visible.dispatchEvent(new Event('change', {bubbles: true}));
                return true;
            }""", note)
            if not filled:
                raise Exception('找不到訂單備註 textarea')
            await page.wait_for_timeout(500)
            await page.locator("button:has-text('記錄備註')").click()
            await page.wait_for_timeout(1500)
            push('訂單備註已記錄', 'ok')

            push(f'全部完成！工單 ID：{ticket_id}', 'ok')
            await browser.close()
            return {'success': True, 'ticket_id': ticket_id, 'use_date': use_date,
                    'ticket_deadline': tdl, 'follow_deadline': fdl,
                    'is_urgent': urgent, 'diff_days': diff, 'order_id': resolved_order_id}

        except StopRequested:
            push('已手動停止', 'error')
            try:
                await browser.close()
            except Exception:
                pass
            return {'success': False, 'error': '已手動停止', 'ticket_id': ticket_id, 'order_id': resolved_order_id}

        except Exception as e:
            push(f'錯誤：{str(e)}', 'error')
            await browser.close()
            return {'success': False, 'error': str(e), 'ticket_id': ticket_id, 'order_id': resolved_order_id}


@app.route('/api/run_general', methods=['POST'])
def api_run_general():
    data = request.json or {}
    order_ids = [o.strip().upper() for o in data.get('order_ids', '').strip().splitlines() if o.strip()]
    supplier_ids = [s.strip() for s in data.get('supplier_ids', '').strip().splitlines() if s.strip()]
    username = data.get('username', '').strip() or BE2_USERNAME
    password = data.get('password', '').strip() or BE2_PASSWORD
    operator = data.get('operator', '').strip() or username.split('@')[0]
    cat_l1 = data.get('cat_l1', '').strip()
    cat_l2 = data.get('cat_l2', '').strip()
    cat_l3 = data.get('cat_l3', '').strip()
    notif_content = data.get('notif_content', '').strip()
    notify_customer = data.get('notify_customer', 'yes').strip()

    if not username or not password:
        return jsonify({'error': '請輸入 BE2 帳號與密碼'}), 400
    if not cat_l1 or not cat_l2 or not cat_l3:
        return jsonify({'error': '請選擇完整工單分類'}), 400
    if not notif_content:
        return jsonify({'error': '請填入後拋備註 / 通知內容'}), 400

    pairs = []
    max_len = max(len(order_ids), len(supplier_ids)) if (order_ids or supplier_ids) else 0
    if max_len == 0:
        return jsonify({'error': '請輸入 KKday 訂單編號或供應商編號'}), 400
    for i in range(max_len):
        oid = order_ids[i] if i < len(order_ids) else ''
        sid = supplier_ids[i] if i < len(supplier_ids) else ''
        pairs.append((oid, sid))

    results = []
    global _waiting_count
    _stop_event.clear()
    for oid, sid in pairs:
        if _stop_event.is_set():
            break
        prog = []
        with _queue_lock:
            _waiting_count += 1
        prog.append({'msg': f'排隊等候中（目前前面有 {_waiting_count - 1} 個）...', 'status': 'info'}) if _waiting_count > 1 else None
        _semaphore.acquire()
        with _queue_lock:
            _waiting_count -= 1
        try:
            r = asyncio.run(run_general_single(oid, sid, cat_l1, cat_l2, cat_l3,
                                               notif_content, prog, username, password, operator, notify_customer))
        finally:
            _semaphore.release()
        resolved_id = r.get('order_id', oid or sid)
        result = {'order_id': resolved_id, 'progress': prog, **r}
        results.append(result)
        save_record({
            'timestamp': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
            'operator': operator,
            'order_id': resolved_id,
            'success': r.get('success', False),
            'ticket_id': r.get('ticket_id', ''),
            'use_date': r.get('use_date', ''),
            'ticket_deadline': r.get('ticket_deadline', ''),
            'follow_deadline': r.get('follow_deadline', ''),
            'is_urgent': r.get('is_urgent', False),
            'error': r.get('error', ''),
            'type': f'一般工單/{cat_l1}/{cat_l2}/{cat_l3}',
        })
    return jsonify({'success': all(r.get('success') for r in results), 'results': results})


@app.route('/api/run_notification', methods=['POST'])
def api_run_notification():
    data = request.json or {}
    order_ids = [o.strip().upper() for o in data.get('order_ids', '').strip().splitlines() if o.strip()]
    supplier_ids = [s.strip() for s in data.get('supplier_ids', '').strip().splitlines() if s.strip()]
    username = data.get('username', '').strip() or BE2_USERNAME
    password = data.get('password', '').strip() or BE2_PASSWORD
    operator = data.get('operator', '').strip() or username.split('@')[0]
    notification_content = data.get('notification_content', '').strip()
    if not username or not password:
        return jsonify({'error': '請輸入 BE2 帳號與密碼'}), 400
    if not notification_content:
        return jsonify({'error': '請輸入行前通知內容'}), 400

    # 合併訂單編號和供應商編號成 pairs
    pairs = []
    max_len = max(len(order_ids), len(supplier_ids)) if (order_ids or supplier_ids) else 0
    if max_len == 0:
        return jsonify({'error': '請輸入 KKday 訂單編號或供應商編號'}), 400
    for i in range(max_len):
        oid = order_ids[i] if i < len(order_ids) else ''
        sid = supplier_ids[i] if i < len(supplier_ids) else ''
        pairs.append((oid, sid))

    results = []
    global _waiting_count
    _stop_event.clear()
    for oid, sid in pairs:
        if _stop_event.is_set():
            break
        prog = []
        with _queue_lock:
            _waiting_count += 1
        prog.append({'msg': f'排隊等候中（目前前面有 {_waiting_count - 1} 個）...', 'status': 'info'}) if _waiting_count > 1 else None
        _semaphore.acquire()
        with _queue_lock:
            _waiting_count -= 1
        try:
            r = asyncio.run(run_notification_flow(oid, sid, notification_content, prog, username, password))
        finally:
            _semaphore.release()
        resolved_id = r.get('order_id', oid or sid)
        result = {'order_id': resolved_id, 'progress': prog, **r}
        results.append(result)
        save_record({
            'timestamp': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
            'operator': operator,
            'order_id': resolved_id,
            'success': r.get('success', False),
            'ticket_id': r.get('ticket_id', ''),
            'use_date': r.get('use_date', ''),
            'ticket_deadline': r.get('ticket_deadline', ''),
            'follow_deadline': r.get('follow_deadline', ''),
            'is_urgent': r.get('is_urgent', False),
            'error': r.get('error', ''),
            'type': '行前通知',
        })
    return jsonify({'success': all(r.get('success') for r in results), 'results': results})


@app.route('/api/stop', methods=['POST'])
def api_stop():
    _stop_event.set()
    return jsonify({'ok': True})


@app.route('/api/export_results', methods=['POST'])
def export_results():
    data = request.json or {}
    results = data.get('results', [])
    title = data.get('title', '執行結果')
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(['訂單編號', '工單 ID', '狀態', '最晚處理', '後拋截止', '急單', '錯誤訊息'])
    for r in results:
        ws.append([
            r.get('order_id', ''),
            '#' + str(r.get('ticket_id', '')) if r.get('ticket_id') else '',
            '成功' if r.get('success') else '失敗',
            r.get('ticket_deadline', '') or '',
            r.get('follow_deadline', '') or '',
            '是' if r.get('is_urgent') else '',
            r.get('error', '') or '',
        ])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=f'{title}.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/notification_template', methods=['GET'])
def notification_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '批次對客通知'
    ws.append(['KKday訂單編號', '對客通知內容 （請輸入中英版本 or 英文）'])
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     download_name='對客通知批次範本.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/parse_notification_excel', methods=['POST'])
def parse_notification_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '請上傳檔案'}), 400
    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            oid = str(row[0]).strip().upper() if row and row[0] else ''
            content = str(row[1]).strip() if row and len(row) > 1 and row[1] else ''
            if oid and content:
                rows.append({'order_id': oid, 'content': content})
    except Exception as e:
        return jsonify({'error': f'Excel 解析失敗：{e}'}), 400
    return jsonify({'count': len(rows), 'rows': rows})


@app.route('/api/run_notification_batch', methods=['POST'])
def api_run_notification_batch():
    file = request.files.get('file')
    username = request.form.get('username', '').strip() or BE2_USERNAME
    password = request.form.get('password', '').strip() or BE2_PASSWORD
    operator = request.form.get('operator', '').strip() or username.split('@')[0]
    if not username or not password:
        return jsonify({'error': '請輸入 BE2 帳號與密碼'}), 400
    if not file:
        return jsonify({'error': '請上傳 Excel 檔案'}), 400
    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            oid = str(row[0]).strip().upper() if row and row[0] else ''
            content = str(row[1]).strip() if row and len(row) > 1 and row[1] else ''
            if oid and content:
                rows.append((oid, content))
    except Exception as e:
        return jsonify({'error': f'Excel 解析失敗：{e}'}), 400
    if not rows:
        return jsonify({'error': 'Excel 內無有效資料（請確認 A 欄訂單編號、B 欄通知內容皆已填入）'}), 400

    results = []
    global _waiting_count
    _stop_event.clear()
    for oid, content in rows:
        if _stop_event.is_set():
            break
        prog = []
        with _queue_lock:
            _waiting_count += 1
        if _waiting_count > 1:
            prog.append({'msg': f'排隊等候中（目前前面有 {_waiting_count - 1} 個）...', 'status': 'info'})
        _semaphore.acquire()
        with _queue_lock:
            _waiting_count -= 1
        try:
            r = asyncio.run(run_notification_flow(oid, '', content, prog, username, password))
        finally:
            _semaphore.release()
        resolved_id = r.get('order_id', oid)
        result = {'order_id': resolved_id, 'progress': prog, **r}
        results.append(result)
        save_record({
            'timestamp': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
            'operator': operator,
            'order_id': resolved_id,
            'success': r.get('success', False),
            'ticket_id': r.get('ticket_id', ''),
            'use_date': r.get('use_date', ''),
            'ticket_deadline': r.get('ticket_deadline', ''),
            'follow_deadline': r.get('follow_deadline', ''),
            'is_urgent': r.get('is_urgent', False),
            'error': r.get('error', ''),
            'type': '行前通知',
        })
    return jsonify({'success': all(r.get('success') for r in results), 'results': results})


@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/run', methods=['POST'])
def api_run():
    data = request.json or {}
    order_ids = [o.strip().upper() for o in data.get('order_ids', '').strip().splitlines() if o.strip()]
    supplier_ids = [s.strip() for s in data.get('supplier_ids', '').strip().splitlines() if s.strip()]
    username = data.get('username', '').strip() or BE2_USERNAME
    password = data.get('password', '').strip() or BE2_PASSWORD
    operator = data.get('operator', '').strip() or username.split('@')[0]
    follow_type = data.get('follow_type', 'page').strip()
    op_dates = data.get('op_dates', '').strip()
    wantan_type = data.get('wantan_type', 'mansatisfied').strip()
    max_len = max(len(order_ids), len(supplier_ids)) if (order_ids or supplier_ids) else 0
    if not max_len:
        return jsonify({'error': '請輸入訂單編號或供應商編號'}), 400
    if not username or not password:
        return jsonify({'error': '請輸入 BE2 帳號與密碼'}), 400
    results = []
    global _waiting_count
    _stop_event.clear()
    for i in range(max_len):
        if _stop_event.is_set():
            break
        oid = order_ids[i] if i < len(order_ids) else ''
        sid = supplier_ids[i] if i < len(supplier_ids) else ''
        prog = []
        with _queue_lock:
            _waiting_count += 1
        prog.append({'msg': f'排隊等候中（目前前面有 {_waiting_count - 1} 個）...', 'status': 'info'}) if _waiting_count > 1 else None
        _semaphore.acquire()
        with _queue_lock:
            _waiting_count -= 1
        try:
            r = asyncio.run(run_flow(oid, prog, username, password, follow_type, op_dates, sid, wantan_type))
        except Exception as e:
            prog.append({'msg': f'未預期錯誤：{e}', 'status': 'error'})
            r = {'success': False, 'error': str(e), 'order_id': oid or sid}
        finally:
            _semaphore.release()
        resolved_id = r.get('order_id', oid or sid)
        result = {'order_id': resolved_id, 'input_id': oid or sid, 'progress': prog, **r}
        results.append(result)
        save_record({
            'timestamp': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
            'operator': operator,
            'order_id': resolved_id,
            'success': r.get('success', False),
            'ticket_id': r.get('ticket_id', ''),
            'use_date': r.get('use_date', ''),
            'ticket_deadline': r.get('ticket_deadline', ''),
            'follow_deadline': r.get('follow_deadline', ''),
            'is_urgent': r.get('is_urgent', False),
            'error': r.get('error', ''),
            'type': f'挽單/{wantan_type}',
        })
    return jsonify({'success': all(r.get('success') for r in results), 'results': results})

@app.route('/api/history')
def api_history():
    return jsonify(load_history())

@app.route('/api/stats')
def api_stats():
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    # offset=0 本週, offset=-1 上週, 以此類推
    offset = int(request.args.get('offset', 0))
    week_monday = now - timedelta(days=now.weekday()) + timedelta(weeks=offset)
    week_sunday = week_monday + timedelta(days=6)
    week_start = week_monday.strftime('%Y-%m-%d')
    week_end = week_sunday.strftime('%Y-%m-%d')
    history = load_history()
    today_records = [r for r in history if r.get('timestamp', '').startswith(today)]
    week_records = [r for r in history if week_start <= r.get('timestamp', '')[:10] <= week_end]
    daily = {}
    for r in week_records:
        d = r.get('timestamp', '')[:10]
        if d not in daily:
            daily[d] = {'total': 0, 'success': 0, 'fail': 0}
        daily[d]['total'] += 1
        if r.get('success'):
            daily[d]['success'] += 1
        else:
            daily[d]['fail'] += 1
    return jsonify({
        'today_total': len(today_records),
        'today_success': sum(1 for r in today_records if r.get('success')),
        'today_fail': sum(1 for r in today_records if not r.get('success')),
        'week_total': len(week_records),
        'week_success': sum(1 for r in week_records if r.get('success')),
        'week_fail': sum(1 for r in week_records if not r.get('success')),
        'week_start': week_start,
        'week_end': week_end,
        'daily': daily,
    })

@app.route('/api/health')
def health():
    return jsonify({'ok': True})

@app.route('/api/queue_status')
def queue_status():
    active = MAX_CONCURRENT - _semaphore._value
    return jsonify({'active': active, 'waiting': _waiting_count, 'max': MAX_CONCURRENT})

if __name__ == '__main__':
    import socket
    def find_free_port(start=5001):
        for p in range(start, start + 20):
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.bind(('', p))
                    return p
            except OSError:
                continue
        return start
    port = int(os.environ.get('PORT', 0)) or find_free_port()
    with open('.port', 'w') as f:
        f.write(str(port))
    app.run(host='0.0.0.0', port=port, debug=False)
