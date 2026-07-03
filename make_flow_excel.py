from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = '工單自動化流程對比'

# ── 顏色 ──
C_HEADER  = '1C3A2E'
C_STEP_W  = 'EAF5EE'   # 額滿挽單 — 綠
C_STEP_U  = 'FEF3E8'   # 未成團挽單 — 橙
C_STEP_N  = 'EAEEf5'   # 對客通知 — 藍
C_STEP_B  = 'F0EAFB'   # 批次對客通知 — 紫
C_STEP_G  = 'EDF5EA'   # 一般工單 — 淺綠
C_NONE    = 'F8F8F8'
C_DIFF    = 'FFF3CD'
C_DIFF_BD = 'F0A500'
C_INPUT   = 'FFF8EE'   # 輸入方式列底色

def fill(h): return PatternFill('solid', fgColor=h)
def bd(style='thin', color='CCCCCC'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def ft(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, name='Arial', size=size)
def al(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── 欄寬 ──
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 26   # 額滿挽單
ws.column_dimensions['D'].width = 26   # 未成團挽單
ws.column_dimensions['E'].width = 26   # 對客通知
ws.column_dimensions['F'].width = 28   # 批次對客通知
ws.column_dimensions['G'].width = 26   # 一般工單
ws.column_dimensions['H'].width = 22   # 差異說明

# ── 標題列 ──
ws.row_dimensions[1].height = 32
for col, h in enumerate(['步驟','階段','額滿挽單','未成團挽單','對客通知（單筆）','批次對客通知','一般工單','差異說明'], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = fill(C_HEADER)
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    c.alignment = al()
    c.border = bd('medium', '0A2318')

# ── 流程資料 (步驟, 階段, 額滿挽單, 未成團挽單, 對客通知, 批次對客通知, 一般工單, 差異說明) ──
rows = [
    # 零、輸入方式
    ('—', '零、輸入方式',
     '手動輸入\n（訂單編號）\n挽單類型：額滿\n可改類型：\n依商品頁面 或 依OP提供',
     '手動輸入\n（訂單編號）\n挽單類型：未成團\n可改類型：\n依OP提供（固定）',
     '手動輸入\n（訂單編號\n+ 通知內容）',
     '上傳 Excel 範本\n（A欄：KKday訂單編號\nB欄：各單通知內容）',
     '手動輸入\n（訂單編號）',
     '批次上傳為獨立入口\n---\n額滿：2種可改類型\n未成團：僅依OP提供'),

    # 一、登入
    (1,  '一、登入',
     '登入 BE2', '登入 BE2', '登入 BE2', '登入 BE2（逐筆執行）', '登入 BE2', ''),

    # 二、查訂單
    (2,  '二、查訂單',
     '（若供應商編號）\n查詢 KKday 訂單編號',
     '（若供應商編號）\n查詢 KKday 訂單編號',
     '依訂單編號處理',
     '依 Excel A欄\n訂單編號逐筆處理',
     '（若供應商編號）\n查詢 KKday 訂單編號', ''),

    # 三、建立工單
    (3,  '三、建立工單',
     '前往客服操作台', '前往客服操作台', '前往客服操作台', '前往客服操作台', '前往客服操作台', ''),

    (4,  '',
     '點擊「新增工單」', '點擊「新增工單」', '點擊「新增工單」', '點擊「新增工單」', '點擊「新增工單」', ''),

    (5,  '',
     '選「有訂單」\n填入訂單編號',
     '選「有訂單」\n填入訂單編號',
     '選「有訂單」\n填入訂單編號',
     '選「有訂單」\n填入訂單編號',
     '選「有訂單」\n填入訂單編號', ''),

    (6,  '',
     '工單分類：\n訂單異動→額滿→挽單',
     '工單分類：\n訂單異動→供應商通知→改期',
     '工單分類：\n供應商自理訊息→供應商通知\n→轉達行前注意事項',
     '工單分類：\n供應商自理訊息→供應商通知\n→轉達行前注意事項',
     '工單分類：\nL1 → L2 → L3（動態選擇）',
     '批次與單筆通知\n分類相同\n---\n額滿 vs 未成團\n工單分類不同'),

    (7,  '',
     '點擊最晚處理時間燈泡\n（自動帶入）',
     '點擊最晚處理時間燈泡\n（自動帶入）',
     '點擊最晚處理時間燈泡\n（自動帶入）',
     '點擊最晚處理時間燈泡\n（自動帶入）',
     '點擊最晚處理時間燈泡\n（自動帶入）', ''),

    (8,  '',
     '確認建立工單', '確認建立工單', '確認建立工單', '確認建立工單', '確認建立工單', ''),

    # 四、開始處理
    (9,  '四、開始處理',
     '點進工單 → 開始處理 → 確認',
     '點進工單 → 開始處理 → 確認',
     '點進工單 → 開始處理 → 確認',
     '點進工單 → 開始處理 → 確認',
     '點進工單 → 開始處理 → 確認', ''),

    # 五、後拋 / 對客通知
    (10, '五、後拋\n/ 對客通知',
     '點擊「後拋」',
     '點擊「後拋」',
     '點擊「對客通知」',
     '點擊「對客通知」',
     '點擊「後拋」',
     '對客通知走\n不同流程'),

    (11, '',
     '選「自動分派」',
     '選「自動分派」',
     '收件人自動帶入 → 下一步',
     '收件人自動帶入 → 下一步',
     '選「自動分派」', ''),

    (12, '',
     '後拋備註填入\n「額滿挽單」',
     '後拋備註填入\n「未成團挽單」',
     '填通知主旨：\n「行前通知 Pre-departure Notice」',
     '填通知主旨：\n「行前通知 Pre-departure Notice」',
     '後拋備註填入\nOP 輸入需請客服協助內容',
     '額滿：固定「額滿挽單」\n未成團：固定「未成團挽單」\n一般：填OP輸入內容'),

    (13, '',
     '點擊後拋截止時間燈泡\n（自動帶入）',
     '點擊後拋截止時間燈泡\n（自動帶入）',
     '填通知內容\n（OP 輸入）',
     '填通知內容\n（Excel B欄各單內容）',
     '點擊後拋截止時間燈泡\n（自動帶入）',
     '批次：內容來自\nExcel B欄'),

    (14, '',
     '此後拋是否需通知旅客 → 是',
     '此後拋是否需通知旅客 → 是',
     '下一步（預覽）→ 確認發送',
     '下一步（預覽）→ 確認發送',
     '此後拋是否需通知旅客？\n是 → 繼續步驟 15–17\n否 → 跳至步驟 18',
     '對客通知\n此處結束'),

    (15, '',
     '誰來發送對客通知\n→ 系統自動發送',
     '誰來發送對客通知\n→ 系統自動發送',
     '—',
     '—',
     '〔是〕誰來發送通知\n→ 後拋處理人',
     '額滿/未成團：系統自動發送\n一般：後拋處理人'),

    (16, '',
     '讀取旅客語系',
     '讀取旅客語系',
     '—', '—',
     '〔是〕填通知主旨\n（自動帶入分類路徑）', ''),

    (17, '',
     '計算通知 D/L\n（後拋截止 - 2 小時）',
     '計算通知 D/L\n（後拋截止 - 2 小時）',
     '—', '—',
     '〔是〕填通知內容\n（OP 輸入）', ''),

    (18, '',
     '填通知主旨\n（自動套模板）',
     '填通知主旨\n（自動套模板）',
     '—', '—',
     '確認後拋\n（是 / 否 共用）',
     '挽單有語系\n自動模板'),

    (19, '',
     '填通知內容\n（依語系自動套模板 + D/L）\n額滿模板\n（依商品頁面 或 依OP提供）',
     '填通知內容\n（依語系自動套模板 + D/L）\n未成團模板\n（依OP提供）',
     '—', '—', '—',
     '額滿/未成團各有\n不同語系模板'),

    (20, '',
     '確認後拋', '確認後拋', '—', '—', '—', ''),

    # 六、關閉工單（對客通知專屬）
    (21, '六、關閉工單\n（對客通知）',
     '—', '—',
     '關閉工單',
     '關閉工單',
     '—',
     '只有對客通知\n需關閉工單'),

    # 七、訂單備註
    (22, '七、訂單備註',
     '填入備註\n（挽單自動化 / 工單ID\n最晚處理 / 後拋截止）',
     '填入備註\n（挽單自動化 / 工單ID\n最晚處理 / 後拋截止）',
     '填入備註\n（對客通知自動化\n工單ID / 最晚處理）',
     '填入備註\n（對客通知自動化\n工單ID / 最晚處理）',
     '填入備註\n（一般工單自動化 / 工單ID\n分類路徑 / 最晚處理 / 後拋截止）',
     '備註內容各有不同'),

    (23, '',
     '記錄備註', '記錄備註', '記錄備註', '記錄備註', '記錄備註', ''),
]

# ── 寫入 ──
for r_idx, (step, stage, w, u, n, b, g, diff) in enumerate(rows, 2):
    is_input_row = (step == '—')
    ws.row_dimensions[r_idx].height = 60 if is_input_row else 50

    for col, (val, bg) in enumerate([
        (step,  'E8F0E8' if is_input_row else 'F4F4F4'),
        (stage, '1C3A2E' if is_input_row else ('EAEAEA' if stage else 'F4F4F4')),
        (w,     C_INPUT if is_input_row else (C_NONE if w == '—' else C_STEP_W)),
        (u,     C_INPUT if is_input_row else (C_NONE if u == '—' else C_STEP_U)),
        (n,     C_INPUT if is_input_row else (C_NONE if n == '—' else C_STEP_N)),
        (b,     C_INPUT if is_input_row else (C_NONE if b == '—' else C_STEP_B)),
        (g,     C_INPUT if is_input_row else (C_NONE if g == '—' else C_STEP_G)),
        (diff,  C_DIFF if diff else ('F0F4F0' if is_input_row else 'FAFAFA')),
    ], 1):
        c = ws.cell(row=r_idx, column=col, value=val)
        c.fill = fill(bg)
        c.alignment = al(h='center' if col == 1 else 'left')

        if is_input_row and col == 2:
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        elif col == 1:
            c.font = ft(bold=True)
        elif col == 2:
            c.font = ft(bold=bool(stage))
        elif val == '—':
            c.font = ft(color='AAAAAA')
        elif col == 8 and diff:
            c.font = Font(bold=True, color='7D4E00', name='Arial', size=9)
        else:
            c.font = ft()

        if col == 8 and diff:
            s = Side(style='medium', color=C_DIFF_BD)
            c.border = Border(left=s, right=s, top=Side(style='thin', color=C_DIFF_BD), bottom=Side(style='thin', color=C_DIFF_BD))
        elif is_input_row:
            c.border = bd('medium', '2D6A4F')
        else:
            c.border = bd()

ws.freeze_panes = 'A2'

out = '/Users/winnie.tseng_1/kkday-ticket-bot/工單自動化流程對比.xlsx'
wb.save(out)
print(f'✅ 已儲存：{out}')
